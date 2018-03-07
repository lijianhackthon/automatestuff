#!/usr/bin/env python3

import sys
import openpyxl, pprint

input_file = sys.argv[1]

wb = openpyxl.load_workbook(input_file)

for name in wb.sheetnames:
    sheet = wb[name]
    for row in range(1, sheet.max_row):
        for column in range(1, sheet.max_column-1):
            print(sheet.cell(row=row, column=column).value, end=" ")
        print("")
